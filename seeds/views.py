

from datetime import datetime, timedelta
import json
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse, reverse_lazy
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from django.contrib.auth import authenticate, login, logout
from .forms import EditProductForm, ProductForm
from .models import Product, UserProfile
from django.contrib.auth.decorators import login_required
from .forms import UserProfileForm
from django.shortcuts import render, redirect
from .models import UserProfile,Order
from .forms import UserProfileForm
from django.db.models import F

def index(request):
    # Retrieve the first 5 products
    productdata = Product.objects.all()[:6]
    
    # Query to get the most ordered product
    most_ordered_product = None
    total_ordered_quantity = 0
    most_ordered_product_count = 0
    all_order_items = OrderItem.objects.all()

    for order_item in all_order_items:
        product_count = all_order_items.filter(product=order_item.product).count()
        if product_count > most_ordered_product_count:
            most_ordered_product_count = product_count
            most_ordered_product = order_item.product
            total_ordered_quantity = most_ordered_product_count * order_item.quantity

    # Check if the most ordered product exists
    if most_ordered_product:
        # Get the image URL and price of the most ordered product
        most_ordered_product_image = most_ordered_product.image.url
        most_ordered_product_price = most_ordered_product.price
    else:
        most_ordered_product_image = None
        most_ordered_product_price = None

    context = {
        'productdata': productdata,
        'most_ordered_product': most_ordered_product,
        'total_ordered_quantity': total_ordered_quantity,
        'most_ordered_product_image': most_ordered_product_image,
        'most_ordered_product_price': most_ordered_product_price,
    }
    
    # Render the index.html template with the provided context
    return render(request, 'index.html', context)

def success(request):
    return render(request, 'order_success.html')

def about(request):
    return render(request, 'about.html')



def photogallery(request):
    return render(request, 'photogallery.html')

def contact(request):
    return render(request, 'contact.html')

def combined_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                if user.is_superuser:
                    # Redirect superuser to admin page
                    return redirect('admin_dashboard')
                else:
                    # Redirect regular user to index page
                    return redirect('index')
        else:
            error_message = 'Invalid credentials. Please try again.'
            return render(request, 'user_login.html', {'form': form, 'error_message': error_message})
    
    else:
        form = AuthenticationForm()

    return render(request, 'user_login.html', {'form': form})

@login_required(login_url=combined_login)
def adminpage(request):
    products_at_reorder_level = Product.objects.filter(quantity__lte=F('reorderlevel'))
    products = Product.objects.all()

    context = {
        'products_at_reorder_level': products_at_reorder_level,
        'productdata': products
    }
    
    return render(request, 'user.html', context)




from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.core.mail import send_mail
from .forms import CustomUserCreationForm

def user_register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Log the user in after registration
            login(request, user)

            # Send registration email
            subject = 'Welcome to Thriveseeds'
            message = f'Hi {user.username},\n\nThank you for registering on Thriveseeds!\n\nYour username: {user.username}\nYour password: {form.cleaned_data["password1"]}'
            from_email = 'deninjose0@gmail.com'
            to_email = user.email
            send_mail(subject, message, from_email, [to_email])

            return redirect('user_login')  # Redirect to login page after registration
    else:
        form = CustomUserCreationForm()
    return render(request, 'user_register.html', {'form': form})




@login_required(login_url=combined_login)
def services(request):
    return render(request, 'services.html')


def user_logout(request):
    logout(request)
    # Redirect to the index page after logout

    return redirect('index')


from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.hashers import make_password
from django.contrib import messages

@login_required
def profile_update(request):
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=request.user)
        password_form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid() and password_form.is_valid():
            user = form.save(commit=False)
            password = password_form.cleaned_data['new_password1']
            if password:
                user.password = make_password(password)  # Change password
            user.save()
            password_form.save()
            messages.success(request, 'Your profile has been updated successfully.')
            return redirect('user_login')  # Redirect to the profile update page
    else:
        form = UserProfileForm(instance=request.user)
        password_form = PasswordChangeForm(request.user)
    return render(request, 'profile_update.html', {'form': form, 'password_form': password_form})

@login_required(login_url=combined_login)
def admin_add(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            # Get the form data including quantity, reorderlevel, and category
            name = form.cleaned_data['name']
            description = form.cleaned_data['description']
            price = form.cleaned_data['price']
            image = form.cleaned_data['image']
            quantity = form.cleaned_data['quantity']
            reorderlevel = form.cleaned_data['reorderlevel']
            category = form.cleaned_data['category']

            # Create a new product instance
            new_product = Product.objects.create(
                name=name,
                description=description,
                price=price,
                image=image,
                quantity=quantity,
                reorderlevel=reorderlevel,
                category=category
            )
            # Save the new product to the database
            new_product.save()

            return redirect('adminpage')  # Redirect after adding a product
    else:
        form = ProductForm()

    return render(request, 'admin_add.html', {'form': form})

from django.shortcuts import render
from django.db.models import Count
from .models import Product

def ourproducts(request):
    # Get all unique categories along with their counts
    categories = Product.objects.values('category').annotate(count=Count('category'))
    
    # Check if category filter is applied
    category = request.GET.get('category')
    if category:
        products = Product.objects.filter(category=category)
    else:
        products = Product.objects.all()
    
    return render(request, 'ourproducts.html', {'productdata': products, 'categories': categories})


from django.urls import reverse


@login_required(login_url=combined_login)
def product_list(request): #admin page product view
    products = Product.objects.all()
    return render(request, 'order_list_and_detail.html', {'product': products})


def edit_product(request, product_id):
    product_to_edit = get_object_or_404(Product, pk=product_id)

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product_to_edit)
        if form.is_valid():
            new_quantity = form.cleaned_data.get('quantity')  # Get the new quantity from the form
            current_quantity = product_to_edit.quantity  # Get the current quantity of the product
            updated_quantity = current_quantity # Calculate the updated quantity

            # Update the product's quantity with the new quantity
            product_to_edit.quantity = updated_quantity
            product_to_edit.save()

            return redirect('adminpage')  # Redirect to the user page
    else:
        form = ProductForm(instance=product_to_edit)

    return render(request, 'edit_product.html', {'form': form, 'product': product_to_edit})

def reorder_level(request, product_id):
    product_to_edit = get_object_or_404(Product, pk=product_id)

    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product_to_edit)
        if form.is_valid():
            form.save()
            return redirect('admin_dashboard')  # Redirect to the admin dashboard page
    else:
        form = ProductForm(instance=product_to_edit)

    return render(request, 'edit_product.html', {'form': form, 'product': product_to_edit})



from django.shortcuts import get_object_or_404, render, redirect
from .models import Product  # Assuming Product model is defined in models.py

def delete_product(request, product_id):
    product_to_delete = get_object_or_404(Product, pk=product_id)

    if request.method == 'POST':
        confirmation = request.POST.get('confirmation', None)
        if confirmation == 'confirmed':
            product_to_delete.delete()  # Delete the product
            return redirect('adminpage')  # Redirect after deletion
        else:
            return redirect('adminpage')  # Redirect without deleting if not confirmed

    return render(request, 'confirmation_page.html', {'product': product_to_delete})

    


def search(request):
    query = request.GET.get('query')
    products = Product.objects.filter(name__icontains=query, is_active=True)
    return render(request, 'search_results.html', {'products': products, 'query': query})

def search_view(request):
    query = request.GET.get('query', '')  # Get the 'query' parameter from the request's GET parameters

    if query:
        # Perform a search using the query, ensuring it's not None
        products = Product.objects.filter(name__icontains=query)
    else:
        # Handle the case where 'query' is None or empty
        products = []

    context = {
        'query': query,
        'products': products,
    }

    return render(request, 'search_results.html', context)

from requests import request

from django.core.mail import send_mail

def checkout_view(request):
    cart_items = Cart.objects.filter(user=request.user)
    total_amount = sum(item.quantity * item.product.price for item in cart_items)

    if request.method == 'POST':
        fullname = request.POST.get('fullname')
        address = request.POST.get('address')
        city = request.POST.get('city')
        postal_code = request.POST.get('postal_code')

        # Construct the redirect URL for order summary
        redirect_url = reverse('order_summary')
        params = {
            'fullname': fullname,
            'address': address,
            'city': city,
            'postal_code': postal_code,
            'total_amount': total_amount,
        }
        redirect_url += '?' + '&'.join([f"{key}={value}" for key, value in params.items()])

        # Send email confirmation to the logged-in user's email address
        subject = 'Order Confirmation - ThriveSEEDS'
        message = f"Dear {fullname},\n\n"
        message += f"Thanks for ordering with ThriveSEEDS.\n\n"
        message += f"For order details, click the following link:\n"
        message += f"Or Please click TRACK ORDER button in your ThriveSEEDS account "
        message += "Best regards,\nThriveseeds Team"

        sender_email = 'your@email.com'  # Replace with your sender email
        recipient_email = request.user.email  # Use the logged-in user's email address
        send_mail(subject, message, sender_email, [recipient_email], html_message=message)

        # Redirect to order summary page
        return HttpResponseRedirect(redirect_url)

    return render(request, 'checkout.html', {'total_amount': total_amount})


from django.shortcuts import render, redirect
from .models import Cart, Order

def order_summary_view(request):
    # Retrieve order information from request parameters
    fullname = request.GET.get('fullname')
    address = request.GET.get('address')
    city = request.GET.get('city')
    postal_code = request.GET.get('postal_code')
    total_amount = request.GET.get('total_amount')

    # Create order instance
    order = Order.objects.create(
        user=request.user,
        fullname=fullname,
        address=address,
        city=city,
        postal_code=postal_code,
        total_amount=total_amount,
    )

    # Retrieve and move cart items to the order
    cart_items = Cart.objects.filter(user=request.user)
    for cart_item in cart_items:
        order.items.create(product=cart_item.product, quantity=cart_item.quantity)
        print(fullname, address, city, postal_code, total_amount)

        cart_item.delete()

    # Prepare context for rendering the order summary page
    context = {
        'fullname': fullname,
        'address': address,
        'city': city,
        'postal_code': postal_code,
        'total_amount': total_amount,
        'cart_products': order.items.all()  # Assuming you have a related_name 'items' for the OrderItem model
    }

    return render(request, 'order_summary.html', context)



from .models import Cart
from django.contrib.auth import get_user_model

User = get_user_model()


from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .models import Product, Cart

@login_required(login_url=combined_login)
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, pk=product_id)

    if request.method == 'POST':
        # Check if the product is already in the cart
        existing_cart_item = Cart.objects.filter(user=request.user, product=product).first()

        # If the product is not in the cart or the quantity is less than 6, allow adding to cart
        if not existing_cart_item or existing_cart_item.quantity < 6:
            # Check if the product quantity is greater than 0
            if product.quantity > 0:
                # If adding a new product to the cart or increasing quantity, create/update cart item
                if existing_cart_item:
                    existing_cart_item.quantity += 1
                    existing_cart_item.save()
                else:
                    cart_item = Cart(user=request.user, product=product, quantity=1)
                    cart_item.save()

                # Decrease product quantity by 1
                product.quantity -= 1
                product.save()

                # Redirect to the cart page or product detail page
                return redirect('cart')  # Redirect to cart page or wherever you manage cart items
            else:
                # Handle out of stock case
                return render(request, 'out_of_stock.html', {'product': product})
        else:
            # Show JavaScript popup for maximum quantity reached
            return HttpResponse('<script>alert("You can only add up to 6 quantities of this product to the cart."); window.history.back();</script>')

    return redirect('product_detail', product_id=product_id)  # Redirect back to product detail page if not POST request



def decrease_to_cart(request, product_id):
    # Get the cart item for the specified product and user
    cart_item = Cart.objects.filter(product_id=product_id, user=request.user).first()
    
    if cart_item:
        # Decrease the quantity by 1
        if cart_item.quantity > 1:
            cart_item.quantity -= 1
            cart_item.save()
        else:
            # If the quantity is already 1, remove the item from the cart
            cart_item.delete()

    return redirect('cart')





def remove_from_cart(request, product_id):
    print('--------')
    cart_items = Cart.objects.filter(product_id=product_id,user=request.user).all()
    print('============')
    print("items",cart_items)
    cart_items.delete()
    return redirect('cart')




from django.shortcuts import render, redirect, get_object_or_404
from .models import Cart, Product

@login_required(login_url='combined_login')
def cart(request):
    upcart_items = Cart.objects.filter(user=request.user)
    
    cart_items = {}
    total_price = 0

    for item in upcart_items:
        if item.product.name in cart_items:
            cart_items[item.product.name]['quantity'] += item.quantity
            cart_items[item.product.name]['total_price'] = cart_items[item.product.name]['price'] * cart_items[item.product.name]['quantity']
        else:
            cart_items[item.product.name] = {
                'id': item.product.id,
                'name': item.product.name,
                'price': item.product.price,
                'quantity': item.quantity,
                'total_price': item.product.price * item.quantity,
                'image': item.product.image
            }
        
        total_price += item.product.price * item.quantity
    
    context = {
        'cart_items': cart_items,
        'total_price': total_price
    }
    
    return render(request, 'cart.html', context)

 
def product_detail(request, product_id):
    product = Product.objects.get(pk=product_id)
    if request.method == 'POST':
        # Handle adding product to cart
        if product.quantity > 0:
            # Decrement quantity and add to cart logic here
            product.quantity -= 1
            product.save()
            # Add to cart logic here
            return redirect('cart')  # Redirect to cart page after adding to cart
        else:
            # Product is out of stock, handle this case as needed
            return render(request, 'product_detail.html', {'product': product, 'message': 'Out of Stock'})
    else:
        return render(request, 'product_detail.html', {'product': product})
#wishlist
@login_required(login_url=combined_login)
def add_to_wishlist(request, product_id):
    if 'wishlist' not in request.session:  
        request.session['wishlist'] = []  

    wishlist = request.session['wishlist'] 
    if product_id not in wishlist:  
        wishlist.append(product_id)
        request.session['wishlist'] = wishlist 
        message='Item added to your wishlist!'
    else:
        message='Item already in wishlist'

    url='/'
    return HttpResponse(pop_message(url,message))
#return HttpRespone
    
@login_required(login_url=combined_login)
def remove_to_wishlist(request, product_id):
    wishlist = request.session['wishlist']   
    wishlist.remove(product_id)
    request.session['wishlist']=wishlist
    return redirect('index')





@login_required(login_url=combined_login)
def view_to_wishlist(request):
    if 'wishlist' not in request.session:  
        context={
            'product':'',
        }
    else:
        wishlist = request.session['wishlist'] 
        data=Product.objects.filter(id__in=wishlist)
        
        context={
            'data':data
        }

    return render(request,'wishlist.html',context)



#custom popup message

def pop_message(url,message):
    url=url
    x=f'''
        <script>
            alert("{message}");
            window.location.href = "{url}"; 
        </script>
    '''
    return(x)



from django.db.models import Sum

@login_required
def admin_order_view(request):
    # Admin view to display all orders with aggregated quantities for each product
    aggregated_orders = Order.objects.values('items__product__name').annotate(total_quantity=Sum('items__quantity'))
    orders = Order.objects.all()
    context = {'orders': orders, 'aggregated_orders': aggregated_orders}
    return render(request, 'order_list_and_detail.html', context)


from django.db.models import Sum

@login_required
def order_detail_view(request, order_id):
    # Admin view to display order details
    order = Order.objects.get(id=order_id)
    
    # Get order items with related product details including image
    items = order.items.select_related('product').values('product__name', 'product__image').annotate(total_quantity=Sum('quantity'))
    
    # Calculate the total quantity of all products
    total_quantity = order.items.aggregate(total=Sum('quantity'))['total']
    
    context = {'order': order, 'items': items, 'total_quantity': total_quantity}
    return render(request, 'order_detail.html', context)


@login_required
def update_status(request, order_id):
    # Update order status view
    if request.method == 'POST':
        new_status = request.POST.get('status')
        try:
            order = Order.objects.get(pk=order_id)
            order.status = new_status
            order.save()
            messages.success(request, 'Order status updated successfully.')
        except Order.DoesNotExist:
            messages.error(request, 'Order not found.')
    return redirect('order_list_and_detail')  # Redirect to orders list

def order_list_and_detail(request):
    orders = Order.objects.all()
    context = {'orders': orders}
    return render(request, 'order_list_and_detail.html', context)

@login_required
def user_order_view(request):
    # Admin view to display all orders
    orders = Order.objects.filter(user=request.user).all()
    context = {'orders': orders}
    return render(request, 'order_list_and_detail.html', context)


def order_list_search(request):
    orders = Order.objects.all()

    # Get search criteria from request parameters
    id_query = request.GET.get('id')
    fullname_query = request.GET.get('fullname')
    status_query = request.GET.get('status')

    # Apply filters based on search criteria
    if id_query:
        id_query_with_code = f'A4B3{id_query}'
        orders = orders.filter(id__icontains=id_query)

    if fullname_query:
        orders = orders.filter(fullname__icontains=fullname_query)

    if status_query:
        orders = orders.filter(status=status_query)

    context = {
        'orders': orders
    }
    return render(request, 'order_list_and_detail.html', context)


@login_required
def admin_dashboard(request):
    last_10_orders = Order.objects.all().order_by('-id')[:10]
    products_at_reorder_level = Product.objects.filter(quantity__lte=F('reorderlevel'))
    products = Product.objects.all()


    context = {
        'last_10_orders': last_10_orders,
        'products_at_reorder_level': products_at_reorder_level,
        'productdata': products,
        
    }

    return render(request, 'admin_dashboard.html', context)

@login_required
def stock_management(request):
    products_at_reorder_level = Product.objects.filter(quantity__lte=F('reorderlevel'))
    return render(request, 'stock_management.html', {'products_at_reorder_level': products_at_reorder_level})



from django.shortcuts import render
from .models import OrderItem, Product

def most_ordered_product_view(request):
    # Query to get the most ordered product
    most_ordered_product = None
    total_ordered_quantity = 0
    most_ordered_product_count = 0
    all_order_items = OrderItem.objects.all()

    for order_item in all_order_items:
        product_count = all_order_items.filter(product=order_item.product).count()
        if product_count > most_ordered_product_count:
            most_ordered_product_count = product_count
            most_ordered_product = order_item.product
            total_ordered_quantity = most_ordered_product_count * order_item.quantity

    context = {
        'most_ordered_product': most_ordered_product,
        'total_ordered_quantity': total_ordered_quantity,
    }
    return render(request, 'purchase_bill.html', context)

#supplier register

from django.shortcuts import render, redirect
from django.contrib.auth.hashers import make_password
from .forms import SupplierForm
from django.core.mail import send_mail
from django.conf import settings

@login_required
def register_supplier(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save(commit=False)
            supplier.password = make_password(form.cleaned_data['password'])  # Hashing password
            supplier.save()

            # Send registration email to the supplier
            subject = 'Registration Successful'
            message = f"Hi {supplier.company_name},\n\nYour registration is successful. Here are your login credentials:\n\nUsername: {supplier.username}\nPassword: {form.cleaned_data['password']}"
            email_from = settings.EMAIL_HOST_USER
            recipient_list = [supplier.email]
            send_mail(subject, message, email_from, recipient_list)

            return redirect('stock_management')  # Redirect to a success page
    else:
        form = SupplierForm()
    return render(request, 'supplier_registration.html', {'form': form})

#supplier login
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from .forms import CustomLoginForm

def custom_login(request):
    if request.method == 'POST':
        form = CustomLoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                # Redirect to a success page or dashboard
                return redirect('supplier_dashboard')
            else:
                # Handle invalid login
                error_message = "Invalid username or password."
                return render(request, 'user_login.html', {'form': form, 'error_message': error_message})
    else:
        form = CustomLoginForm()
    return render(request, 'user_login.html', {'form': form})


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import PurchaseOrder, Supplier

@login_required
def supplier_dashboard(request):
    suppliers = Supplier.objects.all()
    supplier_id = request.GET.get('supplier')
    
    if supplier_id:
        purchase_orders = PurchaseOrder.objects.filter(supplier_id=supplier_id)
    else:
        purchase_orders = PurchaseOrder.objects.all()
    
    context = {
        'suppliers': suppliers,
        'purchase_orders': purchase_orders,
    }
    return render(request, 'supplier_dashboard.html', context)



from django.shortcuts import render
from django.http import JsonResponse
from .models import Product


def proceed_order(request):
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        # Retrieve the product from the database
        try:
            product = Product.objects.get(pk=product_id)
        except Product.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Product does not exist'})

        # Update the quantity of the product
        # Example: Increase the quantity by 1
        product.quantity += 1
        product.save()

        return JsonResponse({'success': True, 'message': 'Quantity updated successfully'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})



from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from .models import Product, PurchaseOrder, Supplier
from .forms import OrderForm

@login_required
def order_product(request, product_id):
    product = Product.objects.get(pk=product_id)
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            quantity = form.cleaned_data['quantity']
            supplier = form.cleaned_data['supplier']  # This will be a Supplier instance
            
            # Check if a similar order already exists for the current user and product
            existing_order = PurchaseOrder.objects.filter(user=request.user, product=product).first()
            if existing_order:
                # Update the quantity of the existing order
                existing_order.quantity += quantity
                existing_order.save()
            else:
                # Create a new order
                PurchaseOrder.objects.create(user=request.user, product=product, quantity=quantity, supplier=supplier)
            
            # Send email to the selected supplier
            send_mail(
                subject=f'New Order for {product.name}',
                message=f'Dear {supplier.company_name},\n\nA new order has been placed for {quantity} units of {product.name}.\n\nThank you!',
                from_email='your_email@example.com',
                recipient_list=[supplier.email],
                fail_silently=False,
            )
            
            return redirect('manage_suppliers')  # Redirect to the purchase orders page
    else:
        form = OrderForm()
    return render(request, 'order_product.html', {'form': form, 'product': product})



from django.db.models import F, Sum
@login_required(login_url=combined_login)

def purchase_orders(request):
    purchase_orders = PurchaseOrder.objects.all()
    total_order_value = purchase_orders.aggregate(total=Sum(F('product__price') * F('quantity')))['total']
    return render(request, 'purchase_orders.html', {'purchase_orders': purchase_orders, 'total_order_value': total_order_value})


# views.py

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Product

@csrf_exempt
@login_required(login_url=combined_login)

def update_product_quantity(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data.get('product_id')
        order_quantity = data.get('order_quantity')
        try:
            product = Product.objects.get(pk=product_id)
            product.quantity += order_quantity
            product.save()
            return JsonResponse({'message': 'Product quantity updated successfully'}, status=200)
        except Product.DoesNotExist:
            return JsonResponse({'error': 'Product not found'}, status=404)
    else:
        return JsonResponse({'error': 'Only POST requests are allowed'}, status=405)




from .models import RestockedProduct

def supplier_edit(request, product_id):
    # Retrieve the product instance
    product = get_object_or_404(Product, pk=product_id)

    if request.method == 'POST':
        form = EditProductForm(request.POST)
        if form.is_valid():
            # Get the edited quantity from the form
            edited_quantity = form.cleaned_data['quantity']
            
            # Calculate the new quantity by adding the edited quantity to the current quantity
            new_quantity = product.quantity + edited_quantity
            product.quantity = new_quantity
            product.save()

            # Create a new RestockedProduct object to store restocked product details
            RestockedProduct.objects.create(product=product, quantity=edited_quantity)

            # Remove the corresponding product from the list of purchase orders
            PurchaseOrder.objects.filter(product_id=product_id).delete()

            return redirect('supplier_dashboard')  # Redirect after editing the product
    else:
        # Pre-fill the form with the current quantity
        form = EditProductForm(initial={'quantity': product.quantity})

    return render(request, 'supplier_edit.html', {'form': form, 'product': product})


from django.shortcuts import render, redirect
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.views import PasswordResetView
from django.contrib import messages

class CustomPasswordResetView(PasswordResetView):
    template_name = 'reset_password.html'
    form_class = PasswordResetForm

    def form_valid(self, form):
        messages.success(self.request, "Password reset email has been sent.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Password reset email could not be sent. Please try again.")
        return super().form_invalid(form)



from .models import Supplier
from .forms import SupplierFormedit
@login_required(login_url=combined_login)

def manage_suppliers(request):
    suppliers = Supplier.objects.all()
    return render(request, 'manage_suppliers.html', {'suppliers': suppliers})


from django.core.mail import send_mail
@login_required(login_url=combined_login)

def edit_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, pk=supplier_id)
    old_email = supplier.email

    if request.method == 'POST':
        form = SupplierFormedit(request.POST, instance=supplier)
        if form.is_valid():
            new_supplier = form.save()
            new_email = new_supplier.email

            # Sending email to the updated email address
            subject = 'Supplier Information Updated'
            message = f"Hello,\n\nYour supplier information has been updated. Thank you for using our service.\n\nRegards,\nThriveseeds.com"
            send_mail(subject, message, 'from@example.com', [new_email])

            # If email is updated, also send email to the old email address
            if old_email != new_email:
                subject = 'Email Address Change Notification'
                message_old = f"Hello,\n\nYour email address has been changed from {old_email} to {new_email}. If you did not request this change, please contact us immediately.\n\nRegards,\nThriveseeds.COM"
                send_mail(subject, message_old, 'from@example.com', [old_email])

            return redirect('manage_suppliers')
    else:
        form = SupplierFormedit(instance=supplier)
    return render(request, 'edit_supplier.html', {'form': form})



from django.shortcuts import render


def purchase_order_list(request):
    # Retrieve all purchase orders
    purchase_orders = PurchaseOrder.objects.all()
    return render(request, 'purchase_order_list.html', {'purchase_orders': purchase_orders})



# views.py

from django.shortcuts import render
from .models import RestockedProduct

def restocked_products(request):
    restocked_products = RestockedProduct.objects.all()
    return render(request, 'restocked_products.html', {'restocked_products': restocked_products})



from django.shortcuts import redirect, get_object_or_404
from .models import Supplier
@login_required(login_url=combined_login)

def delete_supplier(request, supplier_id):
    supplier = get_object_or_404(Supplier, pk=supplier_id)
    supplier.delete()
    return redirect('manage_suppliers')


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import PurchaseOrder, Supplier

@login_required
def select_supplier(request, order_id):
    if request.method == 'POST':
        order = get_object_or_404(PurchaseOrder, id=order_id)
        supplier_id = request.POST.get('supplier_id')
        supplier = get_object_or_404(Supplier, id=supplier_id)
        order.supplier = supplier
        order.save()
        return redirect('supplier_dashboard')  # Change to your dashboard view name
    return redirect('supplier_dashboard')  # Change to your dashboard view name

def weather_dashboard(request):
    """
    Renders the weather visual page.
    """
    return render(request, 'weather_visual.html')





# Configure the Gemini API
# views.py

from django.http import JsonResponse
from django.shortcuts import render
import requests
import re
import google.generativeai as genai

# Configure Gemini API
genai.configure(api_key=("AIzaSyBOQAfb2QRH2u9hyh_bKAFPgtPt7X0diNw"))

def extract_location(query):
    # Adjusted regex pattern for broader query formats
    pattern = r'\b(?:Does|Will|Is|Check|Tell\s+me)?\s*(?:it\s+)?(?:rain|raining)\s*in\s*([\w\s]+?)\s*(?:today)?\b'
    match = re.search(pattern, query, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return None

def query_gemini(weather_data):
    # Create the model configuration
    generation_config = {
        "temperature": 1,
        "top_p": 0.95,
        "top_k": 64,
        "max_output_tokens": 512,  # Adjusted token length for a concise response
        "response_mime_type": "text/plain",
    }

    model = genai.GenerativeModel(
        model_name="gemini-1.5-flash",
        generation_config=generation_config,
    )

    # Formulate a prompt based on weather data
    prompt = (f"The weather in {weather_data['location']} currently is described as {weather_data['description']}. "
              f"The temperature is {weather_data['temperature']}°C, "
              f"with a humidity level of {weather_data['humidity']}% and atmospheric pressure of {weather_data['pressure']} hPa. "
              "Explain if it is raining or not, and provide a brief summary of today's weather conditions.")

    # Start chat session with Gemini
    chat_session = model.start_chat(history=[{"role": "user", "parts": [prompt]}])

    # Get response from Gemini API
    response = chat_session.send_message(prompt)
    return response.text

def get_weather(location):
    api_key = "278129cf208a11de0a09c5a9a490a835"  # Replace with your OpenWeatherMap API key
    url = f"https://api.openweathermap.org/data/2.5/weather?q={location}&appid={api_key}&units=metric"

    try:
        response = requests.get(url)
        data = response.json()
        if data['cod'] != 200:
            return None

        return {
            'temperature': data['main']['temp'],
            'description': data['weather'][0]['description'],
            'humidity': data['main']['humidity'],
            'pressure': data['main']['pressure'],
            'location': location,
        }
    except Exception as e:
        print(f"Error fetching weather data: {e}")
        return None
    
@login_required(login_url=combined_login)

def weather_view(request):
    if request.method == "POST":
        query = request.POST.get('location')
        if not query:
            return render(request, "index.html", {"error": "Query is required."})

        # Extract location from query
        location = extract_location(query)
        
        # Debugging: Print or log the extracted location
        print(f"User query: {query}")
        print(f"Extracted location: {location}")

        if not location:
            return render(request, "index.html", {"error": "Could not extract location from the query."})

        # Fetch weather data
        weather_data = get_weather(location)
        if not weather_data:
            return render(request, "index.html", {"error": f"Could not fetch weather data for {location}."})

        # Use Gemini API to explain the weather
        gemini_response = query_gemini(weather_data)

        # Render template with weather data and Gemini response
        return render(request, "index.html", {
            "weather_data": weather_data,
            "gemini_response": gemini_response,
            "extracted_location": location  # Pass extracted location to the template for debugging
        })
    else:
        return render(request, "index.html")


# views.py

from django.shortcuts import render
@login_required(login_url=combined_login)

def weather_map(request):
    # Replace 'your_api_key' with your actual OpenWeatherMap API key
    api_key = '8fcb95aa406943989846fd4511f34d38'
    
    # Pass the API key to the template context
    context = {'api_key': api_key}
    
    return render(request, 'weather_map.html', context)


# views.py

from django.shortcuts import render
from django.http import JsonResponse
import requests

def weather_forecast(request):
    return render(request, 'weather_forecast.html')

import requests
from django.http import JsonResponse
from django.shortcuts import render

# Replace with your API key
API_KEY = 'AIzaSyDWn7FhpBcLeXw4ygXZd-86c-es1LJiWBA'
API_URL = 'https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash-latest:generateContent'

# Define predefined responses
PREDEFINED_RESPONSES = {
    "Can you tell me what this THRIVEseeds is about?": "Hi, I’m Cropsy! Our e-commerce website THRIVEseeds specializes in selling high-quality crop seeds for various agricultural needs. We offer a wide range of seeds with detailed descriptions, pricing, and weather-based recommendations.",
    "What kinds of crop seeds do you sell?": "Cropsy here! We offer a diverse range of crop seeds including vegetables, fruits, grains, and pulses. You can browse our categories to find specific types of seeds.",
    "Can you give me details about a specific seed?": "Sure thing! Just tell me the name or category of the seed you’re interested in, and I’ll provide you with more details.",
    "How does weather affect the seeds I should buy?": "Great question! Weather plays a crucial role in crop growth. I can help you choose the right seeds based on your local climate conditions using our weather forecasts.",
    "Can you recommend seeds based on the current weather?": "Absolutely! Based on your location and the current weather conditions, I can suggest the best seeds for optimal growth. Just let me know your location.",
    "How do I add items to my cart?": "To add items to your cart, simply select the desired seed, choose the quantity, and click the 'Add to Cart' button.",
    "I want to remove an item from my cart. How do I do that?": "No problem! Go to your cart page, find the item you want to remove, and click the 'Remove' button next to it.",
    "How do I check out?": "To check out, go to your cart, review the items, and click the 'Proceed to Checkout' button. Follow the prompts to enter your shipping information and payment details.",
    "I have a problem with my order. Who should I contact?": "If you have any issues with your order, please contact our customer support team through the contact form on our website or by email at deninjose0@gmail.com.",
    "How can I track my order?": "You can track your order by visiting the 'Order Tracking' section on our website and entering your order number.",
    "How do I create an account?": "To create an account, click on the 'Sign Up' button on the homepage, fill out the required information, and submit the form. You’ll receive a confirmation email to complete the registration.",
    "How can I reset my password?": "If you’ve forgotten your password, go to the 'Login' page and click on 'Forgot Password.' Follow the instructions to reset your password.",
    "What weather conditions should I consider when buying seeds?": "When purchasing seeds, you should consider factors like temperature, humidity, rainfall, and soil conditions. I can provide you with weather forecasts to help you make the right decision.",
    "Can you give me today’s weather forecast?": "Sure! Let me check the current weather conditions for your location. Could you share your city or town?",
    "What is the weather forecast for the next 7 days?": "I can provide you with a 7-day weather forecast for your area. Please visit \"weather dashboard\" after login for get weather forecasting data up to 16 days from now.",
    "How does the weather forecasting feature work?": "THRIVEseeds integrates weather data from reliable sources to help you make informed decisions. The forecasts are updated regularly, and I can provide real-time information for your specific area.",
    "What are the available payment options?": "We accept major credit cards, debit cards, UPI, and net banking. You can choose your preferred option during checkout.",
    "How can I contact customer support?": "You can contact our customer support through the contact form on our website or by emailing us at deninjose0@gmail.com.",
    "What’s your favorite color?": "As much as I’d love to have a favorite color, I’m here to help you with crop seed-related queries! Let me know if you need assistance with any products or weather updates.",
    "Tell me a joke.": "I’m more of a seed and weather expert, but I can certainly help you grow some great crops! Let me know if you need assistance with anything else.",
    "How do I fix my car engine?": "I specialize in crop seeds and weather forecasting, so I might not be able to help with that. However, if you have any questions about our products, I’d be happy to assist!",
    "Can you predict the stock market for me?": "I’m here to provide you with weather forecasts and help with crop seed-related queries. If you’re looking for investment advice, I recommend contacting a financial expert.",
    "How can I grow flowers in space?": "That’s an exciting question! While I can help you grow crops on Earth, space gardening is a bit out of my expertise. Let me know if you need any tips on planting crops here on Earth.",
    "Can you recommend seeds based on the current weather?": "Absolutely! To recommend the best seeds for your area, I need to know your location. Please tell me your city or town.",
    "how are you": "I'm just a chatbot here to assist you with crop seed-related questions. How can I help you today?",
}



def chat(request):
    if request.method == 'POST':
        user_message = request.POST.get('message')

        # Retrieve or initialize conversation history
        conversation_history = request.session.get('conversation_history', [])

        # Add user message to conversation history
        conversation_history.append(f"input: {user_message}")

        # Check if the message matches any predefined response
        bot_reply = PREDEFINED_RESPONSES.get(user_message, None)
        
        if not bot_reply:
            # Define headers and data for the API request
            headers = {
                'Content-Type': 'application/json',
            }
            
            # Prepare context: Use the conversation history
            messages = [{'text': message} for message in conversation_history]
            
            # Prepare data with context (previous conversation)
            data = {
                'contents': [
                    {
                        'parts': messages
                    }
                ]
            }

            # Make the API request
            try:
                response = requests.post(f'{API_URL}?key={API_KEY}', headers=headers, json=data)
                response.raise_for_status()  # Raise an exception for HTTP errors
                
                # Parse the JSON response
                api_response = response.json()
                print("API Response:", api_response)  # For debugging
                
                # Extract the bot reply from the response
                bot_reply = api_response['candidates'][0]['content']['parts'][0]['text']
                
                # Limit the response to a certain number of sentences (e.g., 3)
                bot_reply = '. '.join(bot_reply.split('. ')[:3])  # Limits the response to 3 sentences
                
            except requests.RequestException as e:
                # Handle request errors
                print(f"API request error: {e}")
                bot_reply = 'Sorry, there was an error processing your request.'

        # Add bot response to conversation history
        conversation_history.append(f"output: {bot_reply}")

        # Store updated conversation history in session
        request.session['conversation_history'] = conversation_history

        return JsonResponse({'reply': bot_reply})

    # Render the chat interface if not a POST request
    return render(request, 'base.html')


import json
import traceback
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from .models import Product

# Replace with your OpenWeatherMap API key
OPENWEATHER_API_KEY = '8ba9fb88bc5d4bb5a067230d3a43b56f'

def fetch_weather_data(lat, lon):
    """Fetch weather data for the last 5 days using OpenWeather One Call API."""
    import requests
    from datetime import datetime, timedelta

    url = "https://api.openweathermap.org/data/2.5/onecall/timemachine"
    weather_conditions = []

    try:
        for i in range(1, 6):
            params = {
                'lat': lat,
                'lon': lon,
                'dt': int((datetime.now() - timedelta(days=i)).timestamp()),
                'appid': OPENWEATHER_API_KEY
            }
            response = requests.get(url, params=params)
            response.raise_for_status()  # Raises an exception for HTTP errors
            data = response.json()
            weather_conditions.append(data['current']['weather'][0]['main'])
        
        return weather_conditions

    except requests.RequestException as e:
        print(f"Error fetching weather data: {e}")
        return []
    except Exception as e:
        print(f"Unexpected error: {e}")
        return []

def analyze_weather_conditions(conditions):
    """Analyze the past 5 days of weather conditions to recommend crops."""
    # Simplified logic for demonstration purposes
    condition_counts = {condition: conditions.count(condition) for condition in set(conditions)}
    most_common_condition = max(condition_counts, key=condition_counts.get, default='Normal')
    return most_common_condition

@csrf_exempt
def recommend_products(request):
    """Recommend products based on the last 5 days of weather conditions."""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            latitude = data.get('latitude')
            longitude = data.get('longitude')

            if latitude and longitude:
                # Fetch weather data for the last 5 days
                weather_conditions = fetch_weather_data(latitude, longitude)

                # Analyze the weather conditions
                analyzed_condition = analyze_weather_conditions(weather_conditions)

                # Recommend products based on the analyzed weather condition
                recommended_products = Product.objects.filter(weather_condition=analyzed_condition, is_active=True)
                
                if recommended_products.exists():
                    products_list = ''.join([f'<li>{product.name} - {product.price}</li>' for product in recommended_products])
                    message = f'<h3>We recommend the following products based on recent weather:</h3><ul>{products_list}</ul>'
                else:
                    message = '<h3>No recommendations available for the current weather conditions.</h3>'
                
                return JsonResponse({'message': message})
            else:
                return JsonResponse({'error': 'Latitude or longitude is missing'}, status=400)

        except Exception as e:
            # Log the exception with a full traceback
            print("Internal Server Error: ", str(e))
            traceback.print_exc()
            return JsonResponse({'error': 'An internal server error occurred: {}'.format(str(e))}, status=500)

    # For GET request, return a default HTML page
    return render(request, 'recommendation.html')


from django.shortcuts import render
from .models import Product
from collections import Counter, defaultdict
import matplotlib.pyplot as plt
import io
import base64

def product_visualizations(request):
    # Fetch all products
    orders = Order.objects.all()

    products = Product.objects.all()

    # --- First Visualization: Products per Category ---
    categories = [product.category for product in products]
    category_counts = Counter(categories)

    # Data for the first chart
    category_labels = list(category_counts.keys())
    category_values = list(category_counts.values())

    # Create the first bar chart
    plt.figure(figsize=(6, 4))
    plt.bar(category_labels, category_values, color='skyblue')
    plt.xlabel('Category')
    plt.ylabel('Number of Products')
    plt.title('Products per Category')

    # Convert first chart to PNG
    fig1 = io.BytesIO()
    plt.savefig(fig1, format='png')
    fig1.seek(0)
    plot_url_1 = base64.b64encode(fig1.getvalue()).decode('utf8')

    # --- Second Visualization: Products per Category and Weather Condition ---
    category_weather_counts = defaultdict(lambda: defaultdict(int))
    for product in products:
        category_weather_counts[product.category][product.weather_condition] += 1

    # Prepare data for second chart
    weather_conditions = ['Rainy', 'Hot', 'Snowy', 'Humid', 'Normal']
    categories = list(category_weather_counts.keys())
    data = {condition: [category_weather_counts[category].get(condition, 0) for category in categories] for condition in weather_conditions}

    # Plot stacked bar chart for category vs weather condition
    plt.figure(figsize=(10, 6))
    bottom = [0] * len(categories)

    for condition in weather_conditions:
        plt.bar(categories, data[condition], bottom=bottom, label=condition)
        bottom = [sum(x) for x in zip(bottom, data[condition])]

    plt.xlabel('Category')
    plt.ylabel('Number of Products')
    plt.title('Products per Category and Weather Condition')
    plt.legend(title='Weather Condition')
    plt.xticks(rotation=45)

    # Convert second chart to PNG
    fig2 = io.BytesIO()
    plt.savefig(fig2, format='png')
    fig2.seek(0)
    plot_url_2 = base64.b64encode(fig2.getvalue()).decode('utf8')

    order_statuses = [order.status for order in orders]
    status_counts = Counter(order_statuses)
    status_labels = list(status_counts.keys())
    status_values = list(status_counts.values())

    plt.figure(figsize=(6, 4))
    plt.pie(status_values, labels=status_labels, autopct='%1.1f%%', startangle=90)
    plt.title('Order Status Distribution')
    plt.axis('equal')

    fig3 = io.BytesIO()
    plt.savefig(fig3, format='png')
    fig3.seek(0)
    plot_url_3 = base64.b64encode(fig3.getvalue()).decode('utf8')
    plt.close()

    # --- Fourth Visualization: Orders Per City ---
    order_cities = [order.city for order in orders]
    city_counts = Counter(order_cities)
    city_labels = list(city_counts.keys())
    city_values = list(city_counts.values())

    plt.figure(figsize=(6, 4))
    plt.bar(city_labels, city_values, color='lightcoral')
    plt.xlabel('City')
    plt.ylabel('Number of Orders')
    plt.title('Orders Per City')
    plt.xticks(rotation=45)

    fig4 = io.BytesIO()
    plt.savefig(fig4, format='png')
    fig4.seek(0)
    plot_url_4 = base64.b64encode(fig4.getvalue()).decode('utf8')
    plt.close()

    # Pass all plots to the template
    return render(request, 'product_visualizations.html', {
        'plot_url_1': plot_url_1,
        'plot_url_2': plot_url_2,
        'plot_url_3': plot_url_3,
        'plot_url_4': plot_url_4
    })


from django.shortcuts import render
from django.db.models import Sum, F
from datetime import datetime
from .models import OrderItem
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def monthly_sales_report(request):
    # Get the month from the query parameters, default to the current month and year
    month = request.GET.get('month', datetime.now().strftime('%Y-%m'))
    
    # Split the month and year for query
    year, month = map(int, month.split('-'))

    # Calculate the start and end dates for the selected month
    start_date = datetime(year, month, 1)
    if month < 12:
        end_date = datetime(year, month + 1, 1)
    else:
        end_date = datetime(year + 1, 1, 1)

    # Filter orders within the selected month
    orders = Order.objects.filter(created_at__range=[start_date, end_date])
    
    # Calculate total sales for the month
    total_sales = orders.aggregate(total=Sum('total_amount'))['total'] or 0

    # Calculate sales per product for the month
    order_items = OrderItem.objects.filter(order__in=orders)
    product_sales = order_items.values('product__name').annotate(
        total_quantity=Sum('quantity'),
        total_amount=Sum(F('quantity') * F('product__price'))
    ).order_by('-total_amount')

    if 'export' in request.GET:
        # Create an Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Sales Report {month}-{year}".replace("/", "-")

        # Write the headers
        headers = ['Product', 'Quantity Sold', 'Total Sales Amount']
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

        # Write data to the Excel sheet
        for row_num, item in enumerate(product_sales, 2):
            ws[f"A{row_num}"] = item['product__name']
            ws[f"B{row_num}"] = item['total_quantity']
            ws[f"C{row_num}"] = item['total_amount']

        # Add a row for total sales
        ws[f"A{row_num + 1}"] = 'Total Sales'
        ws[f"C{row_num + 1}"] = total_sales

        # Prepare the response to download the Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Sales_Report_{month}_{year}.xlsx'
        wb.save(response)
        return response

    context = {
        'total_sales': total_sales,
        'month': f"{year}-{month:02d}",
        'product_sales': product_sales,
    }
    return render(request, 'monthly_sales.html', context)

